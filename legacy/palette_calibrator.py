#!/usr/bin/env python3
"""
palette_calibrator.py — Contact sheet pour calibrer les 348 palettes Sanzo Wada.

Pour chaque palette toutes les combinaisons (permutation x miroir) sont visibles
en grille. Cliquer sur une vignette la coche comme validee (reclic = decocher).
Enregistrement automatique dans data/palette_calibration.xlsx.

Raccourcis :  <- / ->  naviguer entre palettes     Espace  nouvelle fractale
"""

import itertools
import json
import tkinter as tk
from datetime import datetime
from pathlib import Path

import numpy as np
from PIL import Image, ImageTk
import openpyxl

import render
import fractal
import iteration

# ── Taille des vignettes ──────────────────────────────────────────────────────
THUMB_W = 220
THUMB_H = int(THUMB_W * 1964 / 3024)   # ~143

N_ITER        = 100
MODE          = "oklab"
SMOOTH        = True
EQUALIZE      = True
CLIP_LIMIT    = 1
TRAP_NORM_MAX = 0.5
TRAP_SIGMA    = 0.5

EXCEL_PATH = Path(__file__).parent / "data" / "palette_calibration.xlsx"
HEADERS    = ["palette_index", "name", "perm_index", "color_order",
              "mirror", "validated_at"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _lighten(hex_color: str, amount: int = 28) -> str:
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return "#{:02x}{:02x}{:02x}".format(
        min(255, r + amount), min(255, g + amount), min(255, b + amount))


def _all_perms(colors: list) -> list:
    return [list(p) for p in itertools.permutations(colors)]


def _hexcolor(rgb) -> str:
    return "#{:02x}{:02x}{:02x}".format(*rgb)


def _make_palette(colors: list) -> list:
    n   = len(colors)
    pos = [i / (n - 1) for i in range(n)] if n > 1 else [0.0]
    return [pos, [list(c) for c in colors]]


def _random_fractal_params() -> tuple:
    gen = fractal.FractalGenerator(THUMB_H, THUMB_W, N_ITER, smooth=SMOOTH)
    c   = gen.pick_interesting_c()
    cx  = float(np.random.normal(0.0, TRAP_SIGMA))
    cy  = float(np.random.normal(0.0, TRAP_SIGMA))
    return c, cx, cy


# ── Excel I/O ─────────────────────────────────────────────────────────────────

def _load_validated() -> set:
    """Charge le fichier Excel -> set de (palette_idx, perm_idx, mirror)."""
    if not EXCEL_PATH.exists():
        return set()
    wb  = openpyxl.load_workbook(EXCEL_PATH)
    ws  = wb.active
    out = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        out.add((int(row[0]), int(row[2]), int(row[4])))
    return out


def _excel_add(idx: int, name: str, perm_index: int,
               color_order: list, mirror: int) -> None:
    EXCEL_PATH.parent.mkdir(exist_ok=True)
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)

    now = datetime.now().isoformat(timespec="seconds")
    for row in ws.iter_rows(min_row=2):
        if (row[0].value == idx
                and row[2].value == perm_index
                and row[4].value == mirror):
            row[1].value = name
            row[3].value = json.dumps(color_order)
            row[5].value = now
            wb.save(EXCEL_PATH)
            return

    ws.append([idx, name, perm_index, json.dumps(color_order), mirror, now])
    wb.save(EXCEL_PATH)


def _excel_remove(idx: int, perm_index: int, mirror: int) -> None:
    if not EXCEL_PATH.exists():
        return
    wb   = openpyxl.load_workbook(EXCEL_PATH)
    ws   = wb.active
    keep = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if (int(row[0]) == idx
                and int(row[2]) == perm_index
                and int(row[4]) == mirror):
            continue
        keep.append(list(row))

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(HEADERS)
    for row in keep:
        ws2.append(row)
    wb2.save(EXCEL_PATH)


# ── Frame scrollable ──────────────────────────────────────────────────────────

class _ScrollFrame(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, **kw)
        bg = kw.get("bg", "#1e1e1e")
        cv = tk.Canvas(self, bg=bg, highlightthickness=0)
        sb = tk.Scrollbar(self, orient="vertical", command=cv.yview)
        cv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(cv, bg=bg)
        wid = cv.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>",
                        lambda _e: cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>",
                lambda e: cv.itemconfig(wid, width=e.width))
        cv.bind_all("<MouseWheel>",
                    lambda e: cv.yview_scroll(int(-1 * e.delta / 120), "units"))


# ── Application ───────────────────────────────────────────────────────────────

class PaletteCalibrator:
    BG      = "#1e1e1e"
    BG_VAL  = "#1b4a1b"
    HL_VAL  = "#4a9a4a"
    BG_NONE = "#272727"
    HL_NONE = "#383838"

    def __init__(self):
        self.palettes  = render.load_sanzo_palettes()
        self.names     = render.load_sanzo_names()
        self.validated = _load_validated()

        self._julia_c = complex(-0.7, 0.27)
        self._trap_cx = 0.0
        self._trap_cy = 0.0

        self._thumb_imgs  = {}
        self._img_lbls    = {}
        self._cell_frames = {}
        self._chk_lbls    = {}

        self.idx    = 0
        self._perms = []

        self.root = tk.Tk()
        self.root.title("Calibration Palettes Sanzo Wada")
        self.root.configure(bg=self.BG)
        self.root.geometry("870x760")
        self._build_ui()
        self._load_palette(0)

        self.root.bind("<Right>", lambda _e: self._next_palette())
        self.root.bind("<Left>",  lambda _e: self._prev_palette())
        self.root.bind("<space>", lambda _e: self._randomize())

    # ── UI ────────────────────────────────────────────────────────────────────

    @staticmethod
    def _btn(parent, text, cmd, bg, fg="white",
             font=("Helvetica", 10), padx=8, pady=3):
        lbl = tk.Label(parent, text=text, bg=bg, fg=fg, font=font,
                       padx=padx, pady=pady, cursor="hand2")
        lbl.bind("<Button-1>", lambda _e: cmd())
        lbl.bind("<Enter>",    lambda _e: lbl.config(bg=_lighten(bg)))
        lbl.bind("<Leave>",    lambda _e: lbl.config(bg=bg))
        return lbl

    def _build_ui(self):
        BG = self.BG

        nav = tk.Frame(self.root, bg=BG)
        nav.pack(fill="x", padx=10, pady=(8, 2))

        self._btn(nav, "  <  ", self._prev_palette, "#333",
                  font=("Helvetica", 14)).pack(side="left", padx=2)
        self._btn(nav, "  >  ", self._next_palette, "#333",
                  font=("Helvetica", 14)).pack(side="left", padx=2)
        self._btn(nav, "  Aleatoire  ", self._randomize, "#4a3a10",
                  fg="#f5d078").pack(side="left", padx=10)

        self._nav_label = tk.Label(nav, text="", bg=BG, fg="#ddd",
                                    font=("Helvetica", 12, "bold"))
        self._nav_label.pack(side="left", padx=8)

        self._prog_label = tk.Label(nav, text="", bg=BG, fg="#666",
                                     font=("Helvetica", 9))
        self._prog_label.pack(side="right")

        # En-tetes colonnes (alignes approximativement sur les cellules)
        hdr = tk.Frame(self.root, bg=BG)
        hdr.pack(anchor="w", padx=10, pady=(2, 0))
        tk.Frame(hdr, bg=BG, width=96).pack(side="left")
        for m in (1, 2, 3):
            tk.Label(hdr, text=f"Miroir {m}", bg=BG, fg="#555",
                     font=("Helvetica", 9),
                     width=(THUMB_W + 14) // 7,
                     anchor="center").pack(side="left", padx=3)

        self._status_var = tk.StringVar(value="")
        tk.Label(self.root, textvariable=self._status_var,
                 bg=BG, fg="#888", font=("Helvetica", 9)).pack(
                     anchor="w", padx=12)

        self._scroll = _ScrollFrame(self.root, bg=BG)
        self._scroll.pack(fill="both", expand=True, padx=10, pady=(0, 8))

    # ── Chargement ────────────────────────────────────────────────────────────

    def _load_palette(self, idx: int):
        self.idx    = idx % len(self.palettes)
        self._perms = _all_perms(self.palettes[self.idx][1])
        c, cx, cy   = _random_fractal_params()
        self._julia_c, self._trap_cx, self._trap_cy = c, cx, cy
        self._status_var.set("Calcul en cours...")
        self.root.update_idletasks()
        self._compute_thumbs()
        self._rebuild_grid()
        self._update_labels()
        self._status_var.set("")

    def _compute_thumbs(self):
        gen  = fractal.FractalGenerator(THUMB_H, THUMB_W, N_ITER, smooth=SMOOTH)
        poly = iteration.Poly(1, 0, self._julia_c)
        V    = gen.generate_julia_trap(
            poly, trap_type=0,
            trap_params=np.array([self._trap_cx, self._trap_cy, 0.0]),
            norm_max=TRAP_NORM_MAX)

        self._thumb_imgs.clear()
        for pi, perm in enumerate(self._perms):
            for mirror in (1, 2, 3):
                r = render.FractalRenderer(
                    _make_palette(perm), mode=MODE, n_iter=N_ITER,
                    repeat=mirror, equalize=EQUALIZE, clip_limit=CLIP_LIMIT)
                arr = r.render(V)
                self._thumb_imgs[(pi, mirror)] = ImageTk.PhotoImage(
                    Image.fromarray(arr))

    def _rebuild_grid(self):
        for w in self._scroll.inner.winfo_children():
            w.destroy()
        self._cell_frames.clear()
        self._img_lbls.clear()
        self._chk_lbls.clear()

        BG = self.BG
        for pi, perm in enumerate(self._perms):
            row = tk.Frame(self._scroll.inner, bg=BG)
            row.pack(fill="x", pady=3, padx=4)

            # Swatches
            sw = tk.Frame(row, bg=BG, width=88)
            sw.pack_propagate(False)
            sw.pack(side="left", padx=(0, 8))
            for rgb in perm:
                tk.Label(sw, width=3, height=3, bg=_hexcolor(rgb)).pack(
                    side="left", padx=1)

            # Cellules mirror 1 / 2 / 3
            for mirror in (1, 2, 3):
                is_val  = (self.idx, pi, mirror) in self.validated
                cell_bg = self.BG_VAL  if is_val else self.BG_NONE
                hl      = self.HL_VAL  if is_val else self.HL_NONE

                cell = tk.Frame(row, bg=cell_bg,
                                highlightthickness=2,
                                highlightbackground=hl,
                                highlightcolor=hl)
                cell.pack(side="left", padx=3)
                self._cell_frames[(pi, mirror)] = cell

                img_lbl = tk.Label(cell,
                                   image=self._thumb_imgs[(pi, mirror)],
                                   bg=cell_bg, cursor="hand2")
                img_lbl.pack(padx=2, pady=(2, 0))
                img_lbl.bind("<Button-1>",
                             lambda e, p=pi, m=mirror: self._toggle(p, m))
                self._img_lbls[(pi, mirror)] = img_lbl

                chk = tk.Label(cell,
                               text="  ok  " if is_val else "      ",
                               bg=cell_bg, fg="#88ff88",
                               font=("Courier", 8, "bold"))
                chk.pack(pady=(0, 2))
                self._chk_lbls[(pi, mirror)] = chk

    # ── Toggle ────────────────────────────────────────────────────────────────

    def _toggle(self, pi: int, mirror: int):
        key    = (self.idx, pi, mirror)
        is_val = key not in self.validated

        if is_val:
            self.validated.add(key)
            _excel_add(self.idx, self.names[self.idx],
                       pi, self._perms[pi], mirror)
        else:
            self.validated.discard(key)
            _excel_remove(self.idx, pi, mirror)

        cell_bg = self.BG_VAL  if is_val else self.BG_NONE
        hl      = self.HL_VAL  if is_val else self.HL_NONE

        cell = self._cell_frames.get((pi, mirror))
        if cell:
            cell.config(highlightbackground=hl, highlightcolor=hl, bg=cell_bg)
        lbl = self._img_lbls.get((pi, mirror))
        if lbl:
            lbl.config(bg=cell_bg)
        chk = self._chk_lbls.get((pi, mirror))
        if chk:
            chk.config(text="  ok  " if is_val else "      ", bg=cell_bg)

        self._update_labels()

    # ── Navigation ────────────────────────────────────────────────────────────

    def _next_palette(self):
        self._load_palette(self.idx + 1)

    def _prev_palette(self):
        self._load_palette(self.idx - 1)

    def _randomize(self):
        c, cx, cy = _random_fractal_params()
        self._julia_c, self._trap_cx, self._trap_cy = c, cx, cy
        self._status_var.set("Calcul en cours...")
        self.root.update_idletasks()
        self._compute_thumbs()
        for (pi, mirror), lbl in self._img_lbls.items():
            lbl.config(image=self._thumb_imgs[(pi, mirror)])
        self._status_var.set("")

    # ── Labels ────────────────────────────────────────────────────────────────

    def _update_labels(self):
        n         = len(self.palettes)
        n_pal_val = sum(1 for (p, _pi, _m) in self.validated if p == self.idx)
        n_total   = len(self.validated)
        coches    = f"{n_pal_val} cochee(s)" if n_pal_val else "aucune cochee"
        self._nav_label.config(
            text=f"{self.idx + 1} / {n}  |  {self.names[self.idx]}")
        self._prog_label.config(
            text=f"{coches} cette palette  |  {n_total} au total")

    # ── Run ───────────────────────────────────────────────────────────────────

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    PaletteCalibrator().run()
