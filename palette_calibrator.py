#!/usr/bin/env python3
"""
palette_calibrator.py — Interface de calibration des 348 palettes Sanzo Wada.

Pour chaque combinaison, teste toutes les permutations d'ordre des couleurs
et choisit le nombre de miroir (1–3). Les choix sont enregistrés dans
data/palette_calibration.xlsx.

Raccourcis clavier :
  Entrée   → Valider & Suivant
  →        → Next (prochaine perm / miroir / palette)
  ←        → Palette précédente
  Space    → Aléatoire (nouvelle fractale, même palette)

Lancement :  myenv/bin/python palette_calibrator.py
"""

import itertools
import json
import tkinter as tk
from datetime import datetime
from pathlib import Path

import numpy as np
from PIL import Image, ImageTk
import openpyxl

import render
import fractal
import iteration

# ── Réglages du rendu de prévisualisation ─────────────────────────────────────
PREVIEW_W     = 580
PREVIEW_H     = int(PREVIEW_W * 1964 / 3024)   # ≈ 377, ratio 3024:1964
N_ITER        = 100
MODE          = "oklab"
SMOOTH        = True
EQUALIZE      = True
CLIP_LIMIT    = 5.5
TRAP_NORM_MAX = 0.5
TRAP_SIGMA    = 0.5   # écart-type N(0, σ) pour le centre du trap

EXCEL_PATH = Path(__file__).parent / "data" / "palette_calibration.xlsx"
HEADERS    = ["palette_index", "name", "perm_index", "color_order", "mirror", "validated_at"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _lighten(hex_color: str, amount: int = 30) -> str:
    """Éclaircit légèrement une couleur hex pour l'effet hover."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return "#{:02x}{:02x}{:02x}".format(
        min(255, r + amount), min(255, g + amount), min(255, b + amount))


def _all_perms(colors: list) -> list[list]:
    return [list(p) for p in itertools.permutations(colors)]


def _hexcolor(rgb) -> str:
    return "#{:02x}{:02x}{:02x}".format(*rgb)


def _make_palette(colors: list) -> list:
    n = len(colors)
    pos = [i / (n - 1) for i in range(n)] if n > 1 else [0.0]
    return [pos, [list(c) for c in colors]]


def _random_fractal_params() -> tuple:
    """Retourne (julia_c, trap_cx, trap_cy) aléatoires."""
    gen = fractal.FractalGenerator(PREVIEW_H, PREVIEW_W, N_ITER, smooth=SMOOTH)
    c  = gen.pick_interesting_c()
    cx = float(np.random.normal(0.0, TRAP_SIGMA))
    cy = float(np.random.normal(0.0, TRAP_SIGMA))
    return c, cx, cy


# ── Excel I/O ─────────────────────────────────────────────────────────────────

def _load_validated() -> dict:
    if not EXCEL_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            result[int(row[0])] = {"perm_index": int(row[2]), "mirror": int(row[4])}
    return result


def _save_entry(idx: int, name: str, perm_index: int,
                color_order: list, mirror: int) -> None:
    EXCEL_PATH.parent.mkdir(exist_ok=True)
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)

    for row in ws.iter_rows(min_row=2):
        if row[0].value == idx:
            row[1].value = name
            row[2].value = perm_index
            row[3].value = json.dumps(color_order)
            row[4].value = mirror
            row[5].value = datetime.now().isoformat(timespec="seconds")
            wb.save(EXCEL_PATH)
            return

    ws.append([idx, name, perm_index, json.dumps(color_order),
               mirror, datetime.now().isoformat(timespec="seconds")])
    wb.save(EXCEL_PATH)


# ── Widget : frame scrollable ─────────────────────────────────────────────────

class _ScrollFrame(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, **kw)
        bg = kw.get("bg", "#1e1e1e")
        self._cv = tk.Canvas(self, bg=bg, highlightthickness=0)
        sb = tk.Scrollbar(self, orient="vertical", command=self._cv.yview)
        self._cv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._cv.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(self._cv, bg=bg)
        self._wid = self._cv.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", lambda _e: self._cv.configure(
            scrollregion=self._cv.bbox("all")))
        self._cv.bind("<Configure>", lambda e: self._cv.itemconfig(
            self._wid, width=e.width))
        self._cv.bind_all("<MouseWheel>",
                          lambda e: self._cv.yview_scroll(
                              int(-1 * (e.delta / 120)), "units"))


# ── Application principale ────────────────────────────────────────────────────

class PaletteCalibrator:
    def __init__(self):
        self.palettes  = render.load_sanzo_palettes()
        self.names     = render.load_sanzo_names()
        self.validated = _load_validated()

        # Paramètres fractale courants (randomisés à chaque nouvelle palette)
        self._julia_c = complex(-0.7, 0.27)
        self._trap_cx = 0.0
        self._trap_cy = 0.0
        self._cache_V = None

        # Première palette non validée
        self.idx = next(
            (i for i in range(len(self.palettes)) if i not in self.validated), 0)

        self.root = tk.Tk()
        self.root.title("Calibration Palettes Sanzo Wada")
        self.root.configure(bg="#1e1e1e")
        self._build_ui()
        self._load_palette(self.idx)

        self.root.bind("<Return>", lambda _e: self._validate())
        self.root.bind("<Right>",  lambda _e: self._next_step())
        self.root.bind("<Left>",   lambda _e: self._prev_palette())
        self.root.bind("<space>",  lambda _e: self._randomize())

    # ── Helpers UI ────────────────────────────────────────────────────────────

    @staticmethod
    def _btn(parent, text, cmd, bg, fg="white", font=("Helvetica", 10), padx=8, pady=3):
        """Label utilisé comme bouton — les couleurs sont toujours respectées sur macOS."""
        lbl = tk.Label(parent, text=text, bg=bg, fg=fg, font=font,
                       padx=padx, pady=pady, cursor="hand2")
        lbl.bind("<Button-1>", lambda _e: cmd())
        lbl.bind("<Enter>",    lambda _e: lbl.config(bg=_lighten(bg)))
        lbl.bind("<Leave>",    lambda _e: lbl.config(bg=bg))
        return lbl

    # ── Construction de l'UI ──────────────────────────────────────────────────

    def _build_ui(self):
        BG = "#1e1e1e"

        # Barre navigation + actions
        nav = tk.Frame(self.root, bg=BG)
        nav.pack(fill="x", padx=10, pady=(8, 0))

        self._btn(nav, "← Palette",       self._prev_palette, "#3a3a3a").pack(side="left", padx=2)
        self._btn(nav, "Palette →",       self._next_palette, "#3a3a3a").pack(side="left", padx=2)
        self._btn(nav, "Next →",          self._next_step,    "#2a5070",
                  font=("Helvetica", 10, "bold")).pack(side="left", padx=6)
        self._btn(nav, "⇄ Aléatoire",    self._randomize,    "#4a3a10",
                  fg="#f5d078").pack(side="left", padx=2)

        self._nav_label = tk.Label(nav, text="", bg=BG, fg="#bbb",
                                    font=("Helvetica", 11))
        self._nav_label.pack(side="left", padx=14)

        self._status_label = tk.Label(nav, text="", bg=BG, fg="#5a5",
                                       font=("Helvetica", 10))
        self._status_label.pack(side="right", padx=10)

        self._btn(nav, "✓  Valider & Suivant", self._validate, "#1e6020",
                  font=("Helvetica", 11, "bold"), padx=12).pack(side="right", padx=6)

        # Barre infos fractale + progression
        info = tk.Frame(self.root, bg=BG)
        info.pack(fill="x", padx=10, pady=(3, 0))

        self._fractal_label = tk.Label(info, text="", bg=BG, fg="#555",
                                        font=("Courier", 9))
        self._fractal_label.pack(side="left")

        self._prog_label = tk.Label(info, text="", bg=BG, fg="#666",
                                     font=("Helvetica", 9))
        self._prog_label.pack(side="right")

        # Corps principal
        body = tk.Frame(self.root, bg=BG)
        body.pack(fill="both", expand=True, padx=10, pady=8)

        # Prévisualisation (gauche)
        pf = tk.Frame(body, bg="#000", bd=1, relief="solid")
        pf.pack(side="left", padx=(0, 10))
        self._canvas_lbl = tk.Label(pf, bg="#000", width=PREVIEW_W, height=PREVIEW_H)
        self._canvas_lbl.pack()

        # Panneau de droite
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="y")

        # Ordre des couleurs (scrollable)
        perm_outer = tk.LabelFrame(right, text="Ordre des couleurs",
                                    bg=BG, fg="#ccc", bd=1,
                                    font=("Helvetica", 10, "bold"))
        perm_outer.pack(fill="both", expand=True, pady=(0, 8))
        self._scroll_frame = _ScrollFrame(perm_outer, bg=BG, height=240, width=240)
        self._scroll_frame.pack(fill="both", expand=True, padx=4, pady=4)
        self._perm_var = tk.IntVar(value=0)

        # Miroir
        mirror_frame = tk.LabelFrame(right, text="Miroir (repeat)",
                                      bg=BG, fg="#ccc", bd=1,
                                      font=("Helvetica", 10, "bold"))
        mirror_frame.pack(fill="x")
        self._mirror_var = tk.IntVar(value=1)
        for v in (1, 2, 3):
            tk.Radiobutton(mirror_frame, text=str(v), variable=self._mirror_var,
                           value=v, bg=BG, fg="white",
                           activebackground=BG, activeforeground="white",
                           selectcolor="#333",
                           command=self._on_setting_change).pack(
                               side="left", padx=14, pady=6)

        # Indicateur de séquence
        self._seq_label = tk.Label(right, text="", bg=BG, fg="#555",
                                    font=("Helvetica", 9))
        self._seq_label.pack(pady=(6, 0), anchor="w")

    # ── Chargement d'une palette ──────────────────────────────────────────────

    def _load_palette(self, idx: int):
        self.idx    = idx % len(self.palettes)
        self._perms = _all_perms(self.palettes[self.idx][1])

        saved = self.validated.get(self.idx)
        self._perm_var.set(saved["perm_index"] if saved else 0)
        self._mirror_var.set(saved["mirror"]    if saved else 1)

        self._rebuild_perm_radios()
        self._update_labels()
        # Nouvelle fractale aléatoire pour chaque nouvelle palette
        self._randomize(refresh=False)
        self._refresh_preview()

    def _rebuild_perm_radios(self):
        for w in self._scroll_frame.inner.winfo_children():
            w.destroy()
        for pi, perm in enumerate(self._perms):
            row = tk.Frame(self._scroll_frame.inner, bg="#1e1e1e")
            row.pack(fill="x", padx=2, pady=1)
            tk.Radiobutton(row, variable=self._perm_var, value=pi,
                           bg="#1e1e1e", activebackground="#1e1e1e",
                           selectcolor="#333",
                           command=self._on_setting_change).pack(side="left")
            for rgb in perm:
                tk.Label(row, width=3, bg=_hexcolor(rgb),
                         relief="flat").pack(side="left", padx=1)

    # ── Rendu ─────────────────────────────────────────────────────────────────

    def _refresh_preview(self):
        if self._cache_V is None:
            gen  = fractal.FractalGenerator(PREVIEW_H, PREVIEW_W, N_ITER, smooth=SMOOTH)
            poly = iteration.Poly(1, 0, self._julia_c)
            self._cache_V = gen.generate_julia_trap(
                poly, trap_type=0,
                trap_params=np.array([self._trap_cx, self._trap_cy, 0.0]),
                norm_max=TRAP_NORM_MAX)

        pi     = self._perm_var.get()
        mirror = self._mirror_var.get()
        renderer = render.FractalRenderer(
            _make_palette(self._perms[pi]), mode=MODE, n_iter=N_ITER,
            repeat=mirror, equalize=EQUALIZE, clip_limit=CLIP_LIMIT)
        arr = renderer.render(self._cache_V)
        self._tk_img = ImageTk.PhotoImage(Image.fromarray(arr))
        self._canvas_lbl.config(image=self._tk_img)
        self._update_seq_label()

    # ── Labels ────────────────────────────────────────────────────────────────

    def _update_labels(self):
        n    = len(self.palettes)
        done = len(self.validated)
        self._nav_label.config(
            text=f"{self.idx + 1} / {n}  ·  {self.names[self.idx]}")
        self._prog_label.config(
            text=f"{done} / {n} validées  ({100 * done // n}%)")
        saved = self.validated.get(self.idx)
        if saved:
            self._status_label.config(
                text=f"✓  perm {saved['perm_index']}  ·  miroir {saved['mirror']}",
                fg="#5a5")
        else:
            self._status_label.config(text="non validé", fg="#666")

    def _update_fractal_label(self):
        c = self._julia_c
        self._fractal_label.config(
            text=f"c = {c.real:+.4f}{c.imag:+.4f}j   "
                 f"trap ({self._trap_cx:+.3f}, {self._trap_cy:+.3f})")

    def _update_seq_label(self):
        pi     = self._perm_var.get()
        mirror = self._mirror_var.get()
        n_p    = len(self._perms)
        step   = pi * 3 + mirror
        total  = n_p * 3
        self._seq_label.config(
            text=f"étape {step}/{total}  ·  perm {pi}/{n_p - 1}  ·  miroir {mirror}")

    # ── Actions ───────────────────────────────────────────────────────────────

    def _on_setting_change(self):
        self._refresh_preview()

    def _randomize(self, refresh: bool = True):
        """Nouvelle fractale aléatoire — garde palette/perm/miroir."""
        c, cx, cy     = _random_fractal_params()
        self._julia_c = c
        self._trap_cx = cx
        self._trap_cy = cy
        self._cache_V = None
        self._update_fractal_label()
        if refresh:
            self._refresh_preview()

    def _next_step(self):
        """Séquence : miroir 1→2→3 → perm suivante → palette suivante."""
        pi     = self._perm_var.get()
        mirror = self._mirror_var.get()
        if mirror < 3:
            self._mirror_var.set(mirror + 1)
            self._refresh_preview()
        elif pi < len(self._perms) - 1:
            self._perm_var.set(pi + 1)
            self._mirror_var.set(1)
            self._refresh_preview()
        else:
            self._load_palette(self.idx + 1)

    def _validate(self):
        pi     = self._perm_var.get()
        mirror = self._mirror_var.get()
        _save_entry(self.idx, self.names[self.idx], pi,
                    self._perms[pi], mirror)
        self.validated[self.idx] = {"perm_index": pi, "mirror": mirror}
        self._update_labels()
        # Prochaine palette non validée
        for offset in range(1, len(self.palettes) + 1):
            candidate = (self.idx + offset) % len(self.palettes)
            if candidate not in self.validated:
                self._load_palette(candidate)
                return
        self._load_palette(self.idx + 1)

    def _next_palette(self):
        self._load_palette(self.idx + 1)

    def _prev_palette(self):
        self._load_palette(self.idx - 1)

    # ── Lancement ─────────────────────────────────────────────────────────────

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    PaletteCalibrator().run()
